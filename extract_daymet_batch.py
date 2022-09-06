#!/usr/bin/python3
#adapted from https://github.com/ornldaac/daymet-single-pixel-batch/blob/master/python/daymet_multiple_extraction.py
#requires file in 'sites.xlsx' format for use with Simple Water Balance Model VBA scripts
#must be placed in directory containing sites.xlsx file with a subdirectory called 'daymet' to download daymet single-pixel data to

from openpyxl import load_workbook
import os
import sys
import requests

start_year = 1980
end_year = 2021
daymet_variables = ['dayl', 'prcp', 'srad', 'swe', 'tmax', 'tmin', 'vp']
daymet_years     = [str(year) for year in range(start_year, end_year + 1)]
daymet_url = r'https://daymet.ornl.gov/single-pixel/api/data?lat={}&lon={}'

sites = load_workbook('sites.xlsx')['sites'] # open worksheet sites from sites.xlsx file

lats = []
lons = []
file_names = []
requested_vars = ",".join(daymet_variables)
requested_years = ",".join(daymet_years)

for row in sites.iter_rows():
    file_names.append(row[0].value)
    lats.append(row[1].value)
    lons.append(row[2].value)


var_str = ''
if requested_vars:
    var_str = "&measuredParams=" + requested_vars

years_str = ''
if requested_years:
    years_str = "&year=" + requested_years
    
num_files_requested = len(lats)
num_downloaded = 0
os.chdir('./daymet')
for i in range(num_files_requested):
    curr_url = daymet_url.format(lats[i], lons[i]) + var_str + years_str
    print("Processing:", curr_url)
    res = requests.get(curr_url)
    if not res.ok:
        print("Could not access the following URL:", curr_url)
    else:
        if file_names[i] == "NULL":
            outFname = res.headers["Content-Disposition"].split("=")[-1]
        else:
            outFname = file_names[i]
        text_str = res.content
        outF = open(outFname, 'wb')
        outF.write(text_str)
        outF.close()
        res.close()
        num_downloaded += 1

print("Finished downloading", num_downloaded, "files.")
